from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
section_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
normal_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', bold=True, size=10)

header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
step1_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
step2_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
step3_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
testing_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
milestone_fill = PatternFill(start_color='E6D5F5', end_color='E6D5F5', fill_type='solid')
section_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
section_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
completed_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# SHEET 1: PROJECT PLAN (Main Timeline)
# ============================================================
ws = wb.active
ws.title = "Project Plan"

ws.merge_cells('A1:J1')
ws['A1'] = 'Seafarer PMS — Full Migration & V2 Architecture Project Plan'
ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:J2')
ws['A2'] = 'Starting Fresh — 4-Week Plan with Testing'
ws['A2'].font = Font(name='Calibri', italic=True, size=11, color='666666')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

headers = ['#', 'Step', 'Module / Task', 'Sub-Tasks', 'Week', 'Days', 'Dependencies', 'Deliverables', 'Testing', 'Status']
col_widths = [5, 8, 28, 55, 8, 7, 22, 35, 30, 12]

row = 4
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

ws.row_dimensions[row].height = 28

# ============================================================
# DATA ROWS
# ============================================================

tasks = [
    # ---- STEP 1: LEGACY IDENTITY RESTRUCTURING ----
    # Section Header
    {'section': True, 'text': 'STEP 1: LEGACY IDENTITY RESTRUCTURING — UUID Migrations for All Modules'},

    # 1.1 Vessels
    {'num': '1.1', 'step': 'Step 1', 'module': 'Vessels — UUID Migration',
     'sub': 'Phase 1: Add vuuid column (TEXT NOT NULL UNIQUE) with backfill using gen_random_uuid()\nPhase 2: Add FK constraints from all child tables referencing vessels to vessels(vuuid)\nPhase 3: Refactor all server & frontend code to use vuuid for lookups\nPhase 4: Convert vessels.id from TEXT to SERIAL PRIMARY KEY',
     'week': 'W1', 'days': '1.5', 'deps': 'None — starting module',
     'deliverables': 'vuuid column live, ~43 child table FKs, all code uses vuuid, id is SERIAL',
     'testing': 'Verify all vessel CRUD, child table inserts, frontend navigation', 'status': 'Not Started'},

    # 1.2 Components
    {'num': '1.2', 'step': 'Step 1', 'module': 'Components — UUID Migration',
     'sub': 'Phase 1: Add cuuid column (TEXT NOT NULL UNIQUE) with backfill\nPhase 2: Add FK constraints from ~15 child tables to components(cuuid)\nPhase 3: Refactor server & frontend code to use cuuid\nPhase 4: Convert components.id from TEXT to SERIAL PRIMARY KEY',
     'week': 'W1', 'days': '1.5', 'deps': 'Vessels migration complete',
     'deliverables': 'cuuid column live, 15 child table FKs, all code uses cuuid, id is SERIAL',
     'testing': 'Verify component CRUD, spare links, documents, running hours refs', 'status': 'Not Started'},

    # 1.3 Jobs
    {'num': '1.3', 'step': 'Step 1', 'module': 'Jobs — UUID Migration',
     'sub': 'Phase 1: Add juuid column (TEXT NOT NULL UNIQUE) with backfill\nPhase 2: Add FK constraints from child tables (job_component_links, work_orders, etc.) to jobs(juuid)\nPhase 3: Refactor server & frontend code to use juuid\nPhase 4: Convert jobs.id from TEXT to SERIAL PRIMARY KEY',
     'week': 'W1', 'days': '1', 'deps': 'Components migration complete',
     'deliverables': 'juuid column live, ~4 child table FKs, all code uses juuid, id is SERIAL',
     'testing': 'Verify job CRUD, job-component links, work order creation', 'status': 'Not Started'},

    # 1.4 Work Orders
    {'num': '1.4', 'step': 'Step 1', 'module': 'Work Orders — UUID Migration',
     'sub': 'Phase 1: Add wouuid column (TEXT NOT NULL UNIQUE) with backfill\nPhase 2: Add FK constraints from 4 child tables (executions, execution_details, etc.) to work_orders(wouuid)\nPhase 3: Refactor server & frontend code to use wouuid\nPhase 4: Convert work_orders.id from TEXT to SERIAL PRIMARY KEY',
     'week': 'W1', 'days': '1', 'deps': 'Jobs migration complete',
     'deliverables': 'wouuid column live, 4 child table FKs, all code uses wouuid, id is SERIAL',
     'testing': 'Verify WO CRUD, execution records, status recalculation, auto-generation', 'status': 'Not Started'},

    # Testing checkpoint
    {'num': '1.T1', 'step': 'Step 1', 'module': 'TESTING CHECKPOINT — Core Modules',
     'sub': 'Integration testing across Vessels, Components, Jobs, Work Orders\nVerify FK constraint integrity across all child tables\nTest bulk import with new UUID-based identities\nValidate auto-generation scheduler with SERIAL ids',
     'week': 'W1', 'days': '0.5', 'deps': 'Tasks 1.1–1.4 complete',
     'deliverables': 'All 4 core modules passing integration tests',
     'testing': 'End-to-end: create vessel → component → job → work order flow', 'status': 'Not Started',
     'fill': testing_fill},

    # 1.5 Running Hours
    {'num': '1.5', 'step': 'Step 1', 'module': 'Running Hours — UUID Migration',
     'sub': 'Phase 1: Add rhuuid column to running_hours_audit & component_running_hours_log\nPhase 2: Add FK constraints to child/related tables\nPhase 3: Refactor server & frontend code to use rhuuid\nPhase 4: Convert id to SERIAL PRIMARY KEY',
     'week': 'W2', 'days': '0.5', 'deps': 'Components migration complete',
     'deliverables': 'rhuuid column live, FK constraints, code refactored, id is SERIAL',
     'testing': 'Verify RH CRUD, delta propagation, counter types (MASTER/INHERITED)', 'status': 'Not Started'},

    # 1.6 Spares
    {'num': '1.6', 'step': 'Step 1', 'module': 'Spares — UUID Migration',
     'sub': 'Phase 1: Add suuid column to spares table with backfill\nPhase 2: Add FK constraints from spare_component_links, spare_location_stock, spares_history, inventory_transactions\nPhase 3: Refactor server & frontend code to use suuid\nPhase 4: Convert spares.id to SERIAL PRIMARY KEY',
     'week': 'W2', 'days': '1', 'deps': 'Components migration complete',
     'deliverables': 'suuid column live, FK constraints on child tables, code refactored, id is SERIAL',
     'testing': 'Verify spares CRUD, component links, location stock, inventory transactions', 'status': 'Not Started'},

    # 1.7 Defects
    {'num': '1.7', 'step': 'Step 1', 'module': 'Defects — UUID Migration',
     'sub': 'Phase 1: Add duuid column to defects table with backfill\nPhase 2: Add FK constraints from defect_actions, defect_attachments, recurring_defect_links\nPhase 3: Refactor server & frontend code to use duuid\nPhase 4: Convert defects.id to SERIAL PRIMARY KEY',
     'week': 'W2', 'days': '1', 'deps': 'Vessels, Components migrations complete',
     'deliverables': 'duuid column live, FK constraints, code refactored, id is SERIAL',
     'testing': 'Verify defect CRUD, actions, attachments, recurring defects, SIRE VIQ', 'status': 'Not Started'},

    # 1.8 Certificates & Surveys
    {'num': '1.8', 'step': 'Step 1', 'module': 'Certificates & Surveys — UUID Migration',
     'sub': 'Phase 1: Add certuuid/survuuid columns to certificates & surveys tables\nPhase 2: Add FK constraints to related tables\nPhase 3: Refactor server & frontend code to use UUID lookups\nPhase 4: Convert id columns to SERIAL PRIMARY KEY',
     'week': 'W2', 'days': '0.5', 'deps': 'Vessels migration complete',
     'deliverables': 'UUID columns live, FK constraints, code refactored, ids are SERIAL',
     'testing': 'Verify cert/survey CRUD, admin module, category/group management', 'status': 'Not Started'},

    # Testing checkpoint 2
    {'num': '1.T2', 'step': 'Step 1', 'module': 'TESTING CHECKPOINT — Extended Modules',
     'sub': 'Integration testing: Running Hours, Spares, Defects, Certificates/Surveys\nVerify cross-module FK integrity (e.g., spare-component links, defect-vessel links)\nRegression testing on core modules (Vessels, Components, Jobs, WOs)',
     'week': 'W2', 'days': '0.5', 'deps': 'Tasks 1.5–1.8 complete',
     'deliverables': 'All 8 modules passing integration tests with UUID identities',
     'testing': 'Cross-module flows: component with spares, WO with defects', 'status': 'Not Started',
     'fill': testing_fill},

    # 1.9 Leaf Node Tables
    {'num': '1.9', 'step': 'Step 1', 'module': 'Leaf Node Tables — UUID + SERIAL',
     'sub': 'Add UUID columns and convert id to SERIAL for all remaining leaf tables:\n- change_request, change_request_attachment, change_request_comment\n- import_history, import_change_log, bulk_import_history, bulk_import_errors\n- audit_log, alert_policies, alert_events, alert_deliveries, alert_config\n- fleet tables, equipment_categories, locations, master_data\n- form_definitions, form_versions, form_version_usage\n- ihm_items, ihm_maintenance_log, makers, maker_list, sfi_details, pms_vessel_settings',
     'week': 'W2', 'days': '1', 'deps': 'All parent module migrations complete',
     'deliverables': 'All leaf tables have UUID columns, all ids are SERIAL auto-increment',
     'testing': 'Verify each leaf table CRUD, FK integrity with parent tables', 'status': 'Not Started'},

    # 1.10 Audit Columns
    {'num': '1.10', 'step': 'Step 1', 'module': 'Add Audit Columns to All Tables',
     'sub': 'Add created_at (TIMESTAMP DEFAULT NOW()), updated_at (TIMESTAMP DEFAULT NOW())\nAdd created_by (TEXT), updated_by (TEXT) to all tables\nUpdate server code to populate audit columns on create/update operations\nAdd database trigger or application logic for auto-updating updated_at',
     'week': 'W2', 'days': '1', 'deps': 'Leaf node tables complete (1.9)',
     'deliverables': 'All tables have 4 audit columns, server code populates them consistently',
     'testing': 'Verify audit columns populated on create/update for each module', 'status': 'Not Started'},

    # Milestone
    {'num': 'M1', 'step': '', 'module': 'MILESTONE: Step 1 Complete — All Legacy Identity Restructured',
     'sub': 'All modules have UUID canonical identifiers\nAll ids are SERIAL auto-increment\nAll audit columns in place\nFull regression test pass',
     'week': 'W2', 'days': '0.5', 'deps': 'All Step 1 tasks',
     'deliverables': 'Fully restructured legacy database with UUID identities and audit trail',
     'testing': 'Full regression testing — all modules end-to-end', 'status': 'Not Started',
     'fill': milestone_fill},

    # ---- STEP 2: V2 ARCHITECTURE ----
    {'section': True, 'text': 'STEP 2: V2 ARCHITECTURE — Build V2 Modules for All Domains'},

    # 2.1 Components + Bulk Upload
    {'num': '2.1', 'step': 'Step 2', 'module': 'V2 — Components + Bulk Upload Module',
     'sub': 'Create server/v2/components/ with repository-service-controller-routes layers\nCreate server/v2/bulk/ for bulk upload functionality\nDefine V2 schemas in shared/v2/components/schema.ts and shared/v2/bulk/schema.ts\nWire frontend URL builders + toggle mechanism\nAll CRUD operations: list, create, read, update, delete components\nBulk import with dry-run, preview, apply workflow',
     'week': 'W3', 'days': '1.5', 'deps': 'Step 1 complete (M1)',
     'deliverables': 'V2 Components + Bulk module fully operational, toggle working',
     'testing': 'Side-by-side comparison: legacy vs V2 returns same data', 'status': 'Not Started'},

    # 2.2 Jobs
    {'num': '2.2', 'step': 'Step 2', 'module': 'V2 — Jobs Module',
     'sub': 'Create server/v2/jobs/ with 4-layer architecture\nRe-export component schemas from V2 components module\nImplement job CRUD, job-component linking, maintenance history\nWire frontend URL builders + toggle',
     'week': 'W3', 'days': '0.5', 'deps': 'V2 Components (2.1)',
     'deliverables': 'V2 Jobs module operational, toggle working',
     'testing': 'Verify job-component links, work order references', 'status': 'Not Started'},

    # 2.3 Work Orders
    {'num': '2.3', 'step': 'Step 2', 'module': 'V2 — Work Orders Module',
     'sub': 'Create server/v2/work-orders/ with 4-layer architecture\nRe-export from components + jobs + spares schemas\nImplement WO CRUD, execution records, status computation\nAuto-generation scheduler, status recalculator\nWire frontend URL builders + toggle',
     'week': 'W3', 'days': '1', 'deps': 'V2 Jobs (2.2)',
     'deliverables': 'V2 Work Orders module operational with automation services',
     'testing': 'Verify WO lifecycle, auto-generation, status recalc', 'status': 'Not Started'},

    # 2.4 Spares + Stores
    {'num': '2.4', 'step': 'Step 2', 'module': 'V2 — Spares + Stores Module',
     'sub': 'Create server/v2/spares/ and server/v2/stores/ with 4-layer architecture\nRe-export component schemas\nImplement spare CRUD, component links, location stock, inventory transactions\nStores module: items, ledger\nWire frontend URL builders + toggle',
     'week': 'W3', 'days': '1', 'deps': 'V2 Components (2.1)',
     'deliverables': 'V2 Spares + Stores modules operational',
     'testing': 'Verify spare-component links, inventory, location stock sync', 'status': 'Not Started'},

    # 2.5 Running Hours
    {'num': '2.5', 'step': 'Step 2', 'module': 'V2 — Running Hours Module',
     'sub': 'Create server/v2/running-hours/ with 4-layer architecture\nRe-export component schemas\nImplement RH CRUD, delta propagation, counter types (MASTER/INHERITED/NOT_RH_DRIVEN)\nWire frontend URL builders + toggle',
     'week': 'W3', 'days': '0.5', 'deps': 'V2 Components (2.1)',
     'deliverables': 'V2 Running Hours module operational',
     'testing': 'Verify counter propagation, audit trail, safety validations', 'status': 'Not Started'},

    # Testing checkpoint — Mid V2
    {'num': '2.T1', 'step': 'Step 2', 'module': 'TESTING CHECKPOINT — Core V2 Modules',
     'sub': 'Side-by-side comparison: legacy vs V2 for core modules\nToggle testing: switch between legacy/V2 and verify same behavior\nCross-module V2 flows: create component → job → WO → execution',
     'week': 'W3', 'days': '0.5', 'deps': 'Tasks 2.1–2.5 complete',
     'deliverables': 'Core V2 modules verified, toggle working',
     'testing': 'Complete V2 core regression suite', 'status': 'Not Started',
     'fill': testing_fill},

    # 2.6 Defects
    {'num': '2.6', 'step': 'Step 2', 'module': 'V2 — Defects Module',
     'sub': 'Create server/v2/defects/ with 4-layer architecture\nImplement defect CRUD, actions, attachments, recurring defects\nSIRE VIQ 7 integration, target date extension workflow\nCondition of Class tracking\nWire frontend URL builders + toggle',
     'week': 'W4', 'days': '1', 'deps': 'V2 Components (2.1)',
     'deliverables': 'V2 Defects module operational with all workflows',
     'testing': 'Verify defect lifecycle, recurring defects, VIQ integration', 'status': 'Not Started'},

    # 2.7 Certificates & Surveys
    {'num': '2.7', 'step': 'Step 2', 'module': 'V2 — Certificates & Surveys Module',
     'sub': 'Create server/v2/certificates/ with 4-layer architecture\nImplement cert/survey CRUD, admin 3-tab interface\nConfigurable categories/groups, prefixed ID formats\nWire frontend URL builders + toggle',
     'week': 'W4', 'days': '0.5', 'deps': 'Step 1 Certs migration complete',
     'deliverables': 'V2 Certificates module operational',
     'testing': 'Verify cert CRUD, category management, survey workflows', 'status': 'Not Started'},

    # 2.8 Admin & Supporting Modules
    {'num': '2.8', 'step': 'Step 2', 'module': 'V2 — Admin, Change Requests, Fleet, Forms',
     'sub': 'Create V2 modules for remaining domains:\n- Admin: bulk import, data purging, Fleet Admin Dashboard\n- Change Requests: workflow with apply-approved-changes\n- Fleet: vessel mapping, component mapping, spare mapping\n- Forms: form definitions, versions, usage\nWire frontend URL builders + toggle for each',
     'week': 'W4', 'days': '1', 'deps': 'Core V2 modules (2.1–2.7)',
     'deliverables': 'All supporting V2 modules operational',
     'testing': 'Verify admin operations, change request workflow, fleet mappings', 'status': 'Not Started'},

    # Testing checkpoint — All V2
    {'num': '2.T2', 'step': 'Step 2', 'module': 'TESTING CHECKPOINT — All V2 Modules',
     'sub': 'Side-by-side comparison: legacy vs V2 for every module\nToggle testing for Defects, Certs, Admin modules\nPerformance comparison: V2 vs legacy response times',
     'week': 'W4', 'days': '0.5', 'deps': 'All Step 2 tasks complete',
     'deliverables': 'All V2 modules verified, toggle working for every module',
     'testing': 'Complete V2 regression suite, performance benchmarks', 'status': 'Not Started',
     'fill': testing_fill},

    # Milestone 2
    {'num': 'M2', 'step': '', 'module': 'MILESTONE: Step 2 Complete — Full V2 Architecture Built',
     'sub': 'Every legacy module has a V2 counterpart\nFrontend toggle works for all modules\nSide-by-side verification passed',
     'week': 'W4', 'days': '0', 'deps': 'All Step 2 tasks + testing',
     'deliverables': 'Complete V2 architecture running alongside legacy',
     'testing': 'Sign-off on V2 parity with legacy', 'status': 'Not Started',
     'fill': milestone_fill},

    # ---- STEP 3: CUTOVER & CLEANUP ----
    {'section': True, 'text': 'STEP 3: CUTOVER — Default to V2, Test, Remove Legacy'},

    # 3.1 Default to V2
    {'num': '3.1', 'step': 'Step 3', 'module': 'Default API to V2',
     'sub': 'Change getApiMode() default from "legacy" to "v2"\nKeep toggle available for rollback\nMonitor for issues\nUpdate all documentation to reflect V2 as default',
     'week': 'W4', 'days': '0.5', 'deps': 'All V2 modules verified (M2)',
     'deliverables': 'V2 is the default API for all users, legacy available via toggle',
     'testing': 'Smoke test all modules in V2 default mode', 'status': 'Not Started'},

    # 3.2 Stabilization & Testing
    {'num': '3.2', 'step': 'Step 3', 'module': 'Stabilization & Full Regression Testing',
     'sub': 'Run full regression test suite with V2 as default\nFix any bugs discovered during stabilization\nPerformance testing under V2\nUser acceptance testing period',
     'week': 'W4', 'days': '1', 'deps': 'V2 set as default (3.1)',
     'deliverables': 'V2 stable with no critical bugs, performance validated',
     'testing': 'Full regression, UAT, performance benchmarks', 'status': 'Not Started',
     'fill': testing_fill},

    # 3.3 Remove Legacy
    {'num': '3.3', 'step': 'Step 3', 'module': 'Remove Legacy Code',
     'sub': 'Remove legacy route registrations from server/routes.ts\nRemove legacy services from server/services/\nRemove toggle mechanism (useApiVersion hook, ComponentApiToggle)\nSimplify *ApiV2.ts files to only return V2 URLs\nClean up shared/schema.ts — remove legacy-only types\nRemove legacy storage interface (server/storage.ts)\nOptional: Rename /technical/api/v2/ to /technical/api/',
     'week': 'W4', 'days': '1', 'deps': 'Stabilization passed (3.2)',
     'deliverables': 'Codebase contains only V2 architecture, no legacy code',
     'testing': 'Full regression on clean V2 codebase', 'status': 'Not Started'},

    # 3.4 Final Testing
    {'num': '3.4', 'step': 'Step 3', 'module': 'Final End-to-End Testing & Sign-Off',
     'sub': 'Complete end-to-end testing of entire application\nVerify all modules: Vessels, Components, Jobs, WOs, RH, Spares, Defects, Certs\nVerify admin operations, bulk import, fleet management\nVerify RBAC for Ship/Office/PMS Admin roles\nDatabase integrity verification\nDocumentation review and update',
     'week': 'W4', 'days': '0.5', 'deps': 'Legacy removal (3.3)',
     'deliverables': 'Production-ready V2-only codebase',
     'testing': 'Complete test suite passing, documentation current', 'status': 'Not Started',
     'fill': testing_fill},

    # Final Milestone
    {'num': 'M3', 'step': '', 'module': 'MILESTONE: Project Complete — V2 Architecture Live',
     'sub': 'All legacy code removed\nV2 architecture is sole codebase\nAll UUID identities in place\nAll audit columns active\nFull test coverage',
     'week': 'W4', 'days': '0', 'deps': 'All tasks complete',
     'deliverables': 'Production-ready Seafarer PMS with V2 architecture',
     'testing': 'Final sign-off', 'status': 'Not Started',
     'fill': milestone_fill},
]

row = 5
for t in tasks:
    if t.get('section'):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=t['text'])
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        for c in range(2, 11):
            ws.cell(row=row, column=c).fill = section_header_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 28
        row += 1
        continue

    fill = t.get('fill')
    if not fill:
        if t.get('step') == 'Step 1':
            fill = step1_fill
        elif t.get('step') == 'Step 2':
            fill = step2_fill
        elif t.get('step') == 'Step 3':
            fill = step3_fill
        else:
            fill = milestone_fill

    values = [t['num'], t['step'], t['module'], t['sub'], t['week'], t['days'], t['deps'], t['deliverables'], t['testing'], t['status']]
    sub_lines = t['sub'].count('\n') + 1
    row_height = max(45, sub_lines * 15)
    ws.row_dimensions[row].height = row_height

    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = bold_font if t['num'].startswith('M') else normal_font
        cell.fill = fill
        cell.alignment = left_align if i in [3, 4, 7, 8, 9] else center_align
        cell.border = thin_border

    row += 1

# ============================================================
# SHEET 2: WEEKLY SUMMARY
# ============================================================
ws2 = wb.create_sheet("Weekly Summary")

ws2.merge_cells('A1:F1')
ws2['A1'] = 'Weekly Summary — Task Distribution'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

headers2 = ['Week', 'Focus Area', 'Key Tasks', 'Days Used', 'Cumulative Days', 'Notes']
col_widths2 = [10, 25, 65, 12, 15, 35]

for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = col_widths2[i-1]

weeks = [
    ['Week 1\n(Days 1-5)', 'Core Module Migrations\n(Vessels → Components → Jobs → WOs)',
     '1.1 Vessels UUID (4 phases) — 1.5 days\n1.2 Components UUID (4 phases) — 1.5 days\n1.3 Jobs UUID (4 phases) — 1 day\n1.4 Work Orders UUID (4 phases) — 1 day\n1.T1 Testing Checkpoint: Core Modules — 0.5 day',
     '5.5', '5.5', 'Start with foundation modules\nEach module follows proven 4-phase pattern\nPattern speeds up after first module'],
    ['Week 2\n(Days 6-10)', 'Extended Module Migrations\n+ Leaf Tables + Audit Columns',
     '1.5 Running Hours UUID — 0.5 day\n1.6 Spares UUID — 1 day\n1.7 Defects UUID — 1 day\n1.8 Certificates/Surveys UUID — 0.5 day\n1.T2 Testing Checkpoint: Extended Modules — 0.5 day\n1.9 Leaf Node Tables UUID + SERIAL — 1 day\n1.10 Audit Columns on All Tables — 1 day\nM1 Step 1 Milestone — 0.5 day',
     '6', '11.5', 'Faster pace — pattern is established\nStep 1 fully complete by end of Week 2'],
    ['Week 3\n(Days 11-15)', 'V2 Architecture Build\n(Core Modules)',
     '2.1 V2 Components + Bulk Upload — 1.5 days\n2.2 V2 Jobs — 0.5 day\n2.3 V2 Work Orders — 1 day\n2.4 V2 Spares + Stores — 1 day\n2.5 V2 Running Hours — 0.5 day\n2.T1 Testing Checkpoint: Core V2 — 0.5 day',
     '5', '16.5', 'V2 modules follow established 4-layer pattern\nIncludes toggle mechanism + frontend wiring'],
    ['Week 4\n(Days 16-20)', 'V2 Extended Modules\n+ Cutover + Legacy Removal',
     '2.6 V2 Defects — 1 day\n2.7 V2 Certificates/Surveys — 0.5 day\n2.8 V2 Admin + Supporting — 1 day\n2.T2 Testing Checkpoint: All V2 — 0.5 day\n3.1 Default to V2 — 0.5 day\n3.2 Stabilization & Full Regression — 1 day\n3.3 Remove Legacy Code — 1 day\n3.4 Final E2E Testing & Sign-off — 0.5 day',
     '6', '22.5', 'Tight schedule — may need buffer days\nV2 modules faster with pattern established\nLegacy removal is final step'],
]

for i, w in enumerate(weeks):
    r = 4 + i
    ws2.row_dimensions[r].height = max(80, w[2].count('\n') * 15)
    fills = [step1_fill, step1_fill, step2_fill if i >= 2 else step1_fill, None, None, None]
    row_fill = [step1_fill, step1_fill, step1_fill, step3_fill][i]
    for j, v in enumerate(w):
        cell = ws2.cell(row=r, column=j+1, value=v)
        cell.font = normal_font
        cell.fill = row_fill
        cell.alignment = left_align if j in [1, 2, 5] else center_align
        cell.border = thin_border

# ============================================================
# SHEET 3: MIGRATION PATTERN REFERENCE
# ============================================================
ws3 = wb.create_sheet("Migration Pattern")

ws3.merge_cells('A1:E1')
ws3['A1'] = '4-Phase UUID Migration Pattern — Applied to Each Module'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

headers3 = ['Phase', 'Name', 'What It Does', 'SQL Pattern', 'Verification']
col_widths3 = [8, 22, 45, 55, 40]

for i, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(i)].width = col_widths3[i-1]

phases = [
    ['1', 'Add UUID Column', 'Add a TEXT NOT NULL UNIQUE column (e.g., vuuid, cuuid)\nBackfill existing rows with gen_random_uuid()\nUse safe pattern: nullable → populate → NOT NULL → UNIQUE',
     'ALTER TABLE <table> ADD COLUMN <uuid> TEXT;\nUPDATE <table> SET <uuid> = gen_random_uuid();\nALTER TABLE <table> ALTER COLUMN <uuid> SET NOT NULL;\nALTER TABLE <table> ADD CONSTRAINT <uuid>_unique UNIQUE (<uuid>);',
     'Column exists, NOT NULL, UNIQUE\nAll rows have UUID values\nNo duplicates'],
    ['2', 'Add FK Constraints', 'Add FOREIGN KEY constraints from child tables to parent(uuid)\nChild tables store UUID from parent.uuid column\nEnsures referential integrity at DB level',
     'ALTER TABLE <child>\nADD CONSTRAINT <fk_name>\nFOREIGN KEY (<ref_col>)\nREFERENCES <parent>(<uuid>);',
     'FK constraints visible in information_schema\nChild inserts fail if parent UUID missing\nCascade behavior works'],
    ['3', 'Refactor Code', 'Update all server lookups to use eq(table.uuid, value)\nUpdate frontend navigation to use uuid\nUpdate API routes to accept uuid params\nRemove id-based lookups everywhere',
     'Server: eq(table.uuid, value) instead of eq(table.id, value)\nFrontend: /module/:uuid instead of /module/:id\nAPI: POST body uses uuid for references',
     'grep finds no old id-based lookups\nFrontend URLs use uuid\nAll CRUD operations work with uuid'],
    ['4', 'Convert id to SERIAL', 'Drop old TEXT PRIMARY KEY\nAdd SERIAL (auto-increment integer) PRIMARY KEY\nRemove any old ID generation logic (e.g., WO-xxx)',
     'ALTER TABLE <table> DROP CONSTRAINT <table>_pkey;\nALTER TABLE <table> DROP COLUMN id;\nALTER TABLE <table> ADD COLUMN id SERIAL PRIMARY KEY;',
     'id column is integer type\nnextval sequence exists\nInserts auto-generate id\nNo old ID generation code remains'],
]

for i, p in enumerate(phases):
    r = 4 + i
    ws3.row_dimensions[r].height = 80
    for j, v in enumerate(p):
        cell = ws3.cell(row=r, column=j+1, value=v)
        cell.font = normal_font
        cell.fill = step1_fill
        cell.alignment = left_align if j >= 2 else center_align
        cell.border = thin_border

# Module application table
r = 10
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws3.cell(row=r, column=1, value='Module-by-Module Application').font = section_font
r += 1

mod_headers = ['Module', 'UUID Column Name', 'Estimated Child Tables', 'Complexity', 'Est. Days']
for i, h in enumerate(mod_headers, 1):
    cell = ws3.cell(row=r, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

modules_data = [
    ['Vessels', 'vuuid', '~43 (largest)', 'High', '2'],
    ['Components', 'cuuid', '~15', 'High', '2'],
    ['Jobs', 'juuid', '~4', 'Medium', '1.5'],
    ['Work Orders', 'wouuid', '4', 'Medium-High', '2'],
    ['Running Hours', 'rhuuid', '~2-3', 'Low-Medium', '1'],
    ['Spares', 'suuid', '~4-5', 'Medium', '1.5'],
    ['Defects', 'duuid', '~3-4', 'Medium', '1.5'],
    ['Certificates/Surveys', 'certuuid/survuuid', '~1-2', 'Low', '1'],
    ['Leaf Node Tables', 'Various', 'N/A', 'Medium (volume)', '2'],
]

for i, m in enumerate(modules_data):
    row_num = r + 1 + i
    for j, v in enumerate(m):
        cell = ws3.cell(row=row_num, column=j+1, value=v)
        cell.font = normal_font
        cell.fill = step1_fill if i < 8 else step2_fill
        cell.alignment = center_align
        cell.border = thin_border

# ============================================================
# SHEET 4: V2 ARCHITECTURE REFERENCE
# ============================================================
ws4 = wb.create_sheet("V2 Architecture")

ws4.merge_cells('A1:F1')
ws4['A1'] = 'V2 Module Architecture — Build Guide'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 30

headers4 = ['Module', 'V2 Path', 'Schema Re-exports From', 'Key Endpoints', 'Dependencies', 'Est. Days']
col_widths4 = [20, 30, 25, 50, 20, 10]

for i, h in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws4.column_dimensions[get_column_letter(i)].width = col_widths4[i-1]

v2_modules = [
    ['Components + Bulk', 'server/v2/components/\nserver/v2/bulk/', 'Root (defines base schemas)',
     'CRUD components, documents, class/regulatory\nBulk import: dry-run, preview, apply', 'None', '2'],
    ['Jobs', 'server/v2/jobs/', 'Components',
     'CRUD jobs, job-component links\nMaintenance history', 'V2 Components', '1'],
    ['Work Orders', 'server/v2/work-orders/', 'Components + Jobs + Spares',
     'CRUD WOs, executions, execution details\nStatus computation, auto-generation\nVessel settings', 'V2 Jobs', '1.5'],
    ['Spares + Stores', 'server/v2/spares/\nserver/v2/stores/', 'Components',
     'CRUD spares, component links, location stock\nInventory transactions\nStores items + ledger', 'V2 Components', '1.5'],
    ['Running Hours', 'server/v2/running-hours/', 'Components',
     'RH audit, component RH log\nDelta propagation, counter types', 'V2 Components', '1'],
    ['Defects', 'server/v2/defects/', 'Components (new)',
     'CRUD defects, actions, attachments\nRecurring defects, SIRE VIQ\nTarget date extensions', 'V2 Components', '1.5'],
    ['Certificates/Surveys', 'server/v2/certificates/', 'Standalone (new)',
     'CRUD certs/surveys\nAdmin categories/groups', 'None', '1'],
    ['Admin + Fleet + CR', 'server/v2/admin/\nserver/v2/fleet/\nserver/v2/change-requests/', 'Multiple',
     'Fleet admin dashboard, vessel mapping\nChange request workflow\nBulk admin operations', 'All V2 modules', '1.5'],
]

for i, m in enumerate(v2_modules):
    r = 4 + i
    ws4.row_dimensions[r].height = 55
    for j, v in enumerate(m):
        cell = ws4.cell(row=r, column=j+1, value=v)
        cell.font = normal_font
        cell.fill = step2_fill
        cell.alignment = left_align if j in [1, 2, 3] else center_align
        cell.border = thin_border

# V2 4-layer reference
r = 14
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws4.cell(row=r, column=1, value='V2 4-Layer Architecture Pattern (Applied to Each Module)').font = section_font
r += 1

layers = [
    ['Layer 1', 'Repository', 'server/v2/<module>/repositories/<name>Repository.ts',
     'Direct Drizzle ORM queries via getDb()', 'Data access only, no business logic', ''],
    ['Layer 2', 'Service', 'server/v2/<module>/services/<name>Service.ts',
     'Business logic, validation, orchestration', 'Calls repositories, enforces rules', ''],
    ['Layer 3', 'Controller', 'server/v2/<module>/controllers/<name>Controller.ts',
     'HTTP request/response handling', 'Parses input, calls services, formats output', ''],
    ['Layer 4', 'Routes', 'server/v2/<module>/routes.ts',
     'Express router with asyncHandler wrapper', 'Maps HTTP methods to controller methods', ''],
]

layer_headers = ['Layer', 'Name', 'File Location', 'Responsibility', 'Rules', '']
for i, h in enumerate(layer_headers, 1):
    cell = ws4.cell(row=r, column=i, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border

for i, l in enumerate(layers):
    row_num = r + 1 + i
    ws4.row_dimensions[row_num].height = 30
    for j, v in enumerate(l):
        cell = ws4.cell(row=row_num, column=j+1, value=v)
        cell.font = normal_font
        cell.fill = step2_fill
        cell.alignment = left_align if j >= 2 else center_align
        cell.border = thin_border

# ============================================================
# SHEET 5: LEGEND & NOTES
# ============================================================
ws5 = wb.create_sheet("Legend & Notes")

ws5.merge_cells('A1:D1')
ws5['A1'] = 'Legend & Project Notes'
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 30
ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 40

ws5.cell(row=3, column=1, value='Color Legend').font = section_font

legend = [
    (step1_fill, 'Step 1 — Legacy Identity Restructuring'),
    (step2_fill, 'Step 2 — V2 Architecture Build'),
    (step3_fill, 'Step 3 — Cutover & Legacy Removal'),
    (testing_fill, 'Testing Checkpoint'),
    (milestone_fill, 'Milestone / Sign-off'),
]

for i, (fill, label) in enumerate(legend):
    r = 4 + i
    cell = ws5.cell(row=r, column=1, value='')
    cell.fill = fill
    cell.border = thin_border
    ws5.cell(row=r, column=2, value=label).font = normal_font

ws5.cell(row=10, column=1, value='Key Project Rules').font = section_font

rules = [
    'All schema changes use Drizzle file-based SQL migrations only (drizzle-kit generate)',
    'One logical change per migration file — never batch unrelated changes',
    'Legacy code stays untouched until Step 3 Phase 4 (legacy removal)',
    'V2 modules are additive — they run alongside legacy with frontend toggle',
    'V2 schemas live in shared/v2/<module>/schema.ts, re-export from parent modules',
    'All V2 modules use getDb() from server/db.ts — never use legacy storage interface',
    'Frontend uses URL builder files (*ApiV2.ts) — no hardcoded API paths',
    'UUID columns use TEXT NOT NULL UNIQUE — not actual UUID type',
    'Each module follows the proven 4-phase pattern: UUID → FK → Code Refactor → SERIAL',
    'Audit columns (created_at, updated_at, created_by, updated_by) added after all identity migrations',
    'Testing checkpoints are mandatory between major phases — never skip them',
    'Total estimated timeline: 4 weeks (20 working days) with buffer for overruns',
]

for i, rule in enumerate(rules):
    ws5.cell(row=11 + i, column=1, value=f'{i+1}.').font = bold_font
    ws5.cell(row=11 + i, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws5.cell(row=11 + i, column=2, value=rule).font = normal_font
    ws5.cell(row=11 + i, column=2).alignment = left_align

ws5.cell(row=25, column=1, value='Timeline Summary').font = section_font

summary = [
    ['Step 1: Legacy Identity Restructuring', '~11.5 days', 'Weeks 1-2'],
    ['Step 2: V2 Architecture Build', '~8 days', 'Weeks 3-4'],
    ['Step 3: Cutover & Cleanup', '~3 days', 'Week 4'],
    ['Testing Checkpoints (included in above)', '~3 days', 'Spread across all weeks'],
    ['TOTAL', '~20 working days', '4 weeks'],
]

for i, s in enumerate(summary):
    r = 26 + i
    for j, v in enumerate(s):
        cell = ws5.cell(row=r, column=j+1, value=v)
        cell.font = bold_font if i == 4 else normal_font
        cell.border = thin_border
        if i == 4:
            cell.fill = milestone_fill

# Save
output_path = '/home/runner/workspace/Seafarer_PMS_Migration_Project_Plan.xlsx'
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
